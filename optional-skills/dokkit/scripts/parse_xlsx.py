#!/usr/bin/env python3
"""Parse XLSX files into Dokkit's dual-file format (Markdown + JSON sidecar).

Usage:
    python parse_xlsx.py <input.xlsx>

Output:
    JSON to stdout with 'content_md' and 'metadata' fields.

Requires:
    pip install openpyxl
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({
        "error": "openpyxl not installed. Run: pip install openpyxl"
    }))
    sys.exit(1)


def parse_xlsx(file_path: str) -> dict:
    """Parse an XLSX file and return content + metadata."""
    path = Path(file_path)
    wb = openpyxl.load_workbook(path, data_only=True)

    sections = []
    all_content = []
    key_value_pairs = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sections.append(sheet_name)
        all_content.append(f"## {sheet_name}\n")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            all_content.append("*(empty sheet)*\n")
            continue

        # Detect if first row is a header
        headers = [str(c) if c is not None else "" for c in rows[0]]

        # Build markdown table
        all_content.append("| " + " | ".join(headers) + " |")
        all_content.append("| " + " | ".join(["---"] * len(headers)) + " |")

        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            all_content.append("| " + " | ".join(cells) + " |")

            # Extract key-value pairs from 2-column patterns
            if len(cells) >= 2 and cells[0] and cells[1]:
                # If first column looks like a label (short text, no numbers)
                label = cells[0].strip()
                value = cells[1].strip()
                if len(label) < 50 and not label.replace(" ", "").isdigit():
                    key_value_pairs[label] = value

        all_content.append("")

    content_md = f"# {path.stem}\n\n" + "\n".join(all_content)

    return {
        "content_md": content_md,
        "metadata": {
            "file_name": path.name,
            "file_type": "xlsx",
            "parse_date": datetime.now().isoformat(),
            "key_value_pairs": key_value_pairs,
            "sections": sections,
            "sheet_count": len(wb.sheetnames),
        }
    }


def main():
    if len(sys.argv) != 2:
        print("Usage: python parse_xlsx.py <input.xlsx>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    if not Path(file_path).exists():
        print(json.dumps({"error": f"File not found: {file_path}"}))
        sys.exit(1)

    result = parse_xlsx(file_path)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
